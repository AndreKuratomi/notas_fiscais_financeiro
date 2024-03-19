import pandas as pd
import time

from openpyxl import load_workbook
from pathlib import Path
from tqdm import tqdm

from management_before_django.table_managements.modules.openpyxl_module import adicionar_coluna_contatos, adicionar_coluna_status, contatos_teste, workbook_para_pandas
from management_before_django.table_managements.modules.take_path_from_directory import paths_with_file_name, paths_com_muitos_nomes_de_arquivos

from utils.variables.envs import sheet, sheet_contacts

import ipdb


def filter_table_column(raw_path: Path, edited_path: Path, sheet: str) -> pd.DataFrame:
    """Receives the tables' path, filters it as necessary and inserts it to Pandas dataframe"""

    # PLANILHAS RECÉM-BAIXADAS:
    # Paths:
    (contatos, complete_file_path_to_raw, file_path_to_raw) = paths_com_muitos_nomes_de_arquivos(raw_path)

    # Workbooks:
    workbook_contacts_data = load_workbook(data_only=True, filename=contatos)
    contacts_data = workbook_contacts_data[sheet_contacts]

    workbook_all_data = load_workbook(data_only=True, filename=file_path_to_raw)
    all_data = workbook_all_data[sheet]

    # Funções Openpyxl:
    adicionar_coluna_contatos(all_data, contacts_data, workbook_all_data, complete_file_path_to_raw)
    adicionar_coluna_status(all_data, edited_path, file_path_to_raw, workbook_all_data, sheet)


    # PLANILHA EDITADA:

    (complete_file_path_to_edited, file_path_to_edited) = paths_with_file_name(edited_path)

    workbook_edited = load_workbook(data_only=True, filename=file_path_to_edited)
    table_sheet = workbook_edited[sheet]

    contatos_teste(table_sheet, workbook_edited, complete_file_path_to_edited)
    pandas_dataframe = workbook_para_pandas(table_sheet)

    return pandas_dataframe
